from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT = Path("outputs/019e3085-85ab-7051-975a-670f12230d36/amazon_fba_reimbursement_workflow_example.xlsx")
AS_OF = date(2026, 5, 16)


COLORS = {
    "navy": "17324D",
    "blue": "2F75B5",
    "teal": "0F6B6E",
    "green": "2E7D32",
    "red": "B23A48",
    "orange": "C87900",
    "yellow": "FFF2CC",
    "gray": "F3F6F8",
    "light_blue": "D9EAF7",
    "light_teal": "DDEFEF",
    "light_green": "E2F0D9",
    "light_red": "FCE4D6",
    "white": "FFFFFF",
    "dark_text": "1F2933",
    "medium_text": "52616B",
    "border": "C9D3DC",
}


source_reports = [
    ["Must", "SP-API Reports", "Inventory Ledger Detail", "GET_LEDGER_DETAIL_VIEW_DATA", "Daily", "FNSKU, MSKU, ASIN, ReferenceID, Date", "Primary truth for lost, found, damaged, disposed, returned, removed, and unreconciled units.", "Use eventType=Adjustments and full ledger extracts for reconciliation."],
    ["Must", "SP-API Reports", "FBA Reimbursements", "GET_FBA_REIMBURSEMENTS_DATA", "Daily", "reimbursement-id, case-id, order-id, sku, fnsku, asin", "Prevents duplicate cases and confirms cash/inventory reimbursements, reversals, reasons, and case IDs.", "Pull before every case generation step."],
    ["Must", "SP-API Reports", "FBA Customer Returns", "GET_FBA_FULFILLMENT_CUSTOMER_RETURNS_DATA", "Daily", "order-id, sku, fnsku, return-date, LPN", "Detects refunded items that were returned, missing, damaged, or dispositioned incorrectly.", "Includes return reason, detailed disposition, status, and customer comments."],
    ["Must", "SP-API Reports", "FBA Replacements", "GET_FBA_FULFILLMENT_CUSTOMER_SHIPMENT_REPLACEMENT_DATA", "Daily", "original-order-id, replacement-order-id, sku, asin", "Finds replacements where the original item was not returned or not reimbursed.", "North America and India availability per SP-API docs."],
    ["Must", "SP-API Reports", "Amazon Fulfilled Shipments", "GET_AMAZON_FULFILLED_SHIPMENTS_DATA_GENERAL", "Daily", "amazon-order-id, sku, shipment-date", "Provides original FBA order context for refunds, replacements, and post-order loss claims.", "Use with finances/refund events for post-order cases."],
    ["Must", "Finances API", "Financial Events", "GET /finances/v0/financialEvents", "Daily", "order-id, posted-date, event type", "Confirms refunds, reimbursements, fee charges, and settlement posting before marking recovered.", "Orders from the last 48 hours might lag."],
    ["Must", "Reports API", "Settlement Report V2", "GET_V2_SETTLEMENT_REPORT_DATA_FLAT_FILE_V2", "Per settlement", "settlement-id, order-id, adjustment-id, sku", "Accounting-grade backup for reimbursements, reversals, refunds, fees, and payment reconciliation.", "Automatically scheduled by Amazon; retrieve with getReports."],
    ["Must", "SP-API Reports", "Removal Order Detail", "GET_FBA_FULFILLMENT_REMOVAL_ORDER_DETAIL_DATA", "Daily", "removal order-id, sku, fnsku", "Reconciles requested, cancelled, disposed, shipped, and in-process removal quantities.", "Use with removal shipment detail."],
    ["Must", "SP-API Reports", "Removal Shipment Detail", "GET_FBA_FULFILLMENT_REMOVAL_SHIPMENT_DETAIL_DATA", "Daily", "removal order-id, tracking-number, sku", "Detects lost removal shipments and confirms delivered removals.", "Does not include canceled returns or disposed items."],
    ["Should", "SP-API Reports", "FBA Fee Preview", "GET_FBA_ESTIMATED_FBA_FEES_TXT_DATA", "Weekly", "sku, fnsku, asin", "Baseline for fulfillment fee overcharge checks.", "Compare with charged fees and product dimensions/weight."],
    ["Should", "SP-API Reports", "FBA Storage Fees", "GET_FBA_STORAGE_FEE_CHARGES_DATA", "Monthly", "asin, fnsku, fulfillment_center", "Storage, overage, and dimension-driven fee validation.", "Monthly report; build slower cadence."],
    ["Should", "SP-API Reports", "Long-Term Storage Fee Charges", "GET_FBA_FULFILLMENT_LONGTERM_STORAGE_FEE_CHARGES_DATA", "Monthly", "snapshot-date, sku, fnsku, asin", "Validates aged inventory fees and quantities charged.", "Good for fee-overcharge audit, not classic lost/damaged claims."],
    ["Should", "SP-API Reports", "Inbound Performance", "GET_FBA_FULFILLMENT_INBOUND_NONCOMPLIANCE_DATA", "Daily", "fba-shipment-id, sku, fnsku", "Captures inbound problems, coaching events, and problem quantities.", "Separate from reimbursement eligibility but useful for evidence and root cause."],
    ["Should", "SP-API Reports", "FBA Manage Inventory", "GET_FBA_MYI_UNSUPPRESSED_INVENTORY_DATA", "Daily", "sku, fnsku, asin", "Current available, unsellable, reserved, total, inbound, and researching quantities.", "Useful sanity check before creating lost/damaged cases."],
    ["Should", "Seller Data", "Manufacturing Cost / COGS", "Seller-maintained table", "On change", "sku, fnsku, asin", "Valuation reference for pre-order lost/damaged claims after 2025 policy change.", "Keep invoices and approved cost proof linked."],
    ["Should", "Seller Docs", "Invoices, Packing Slips, POD/BOL", "File storage", "Per shipment", "shipment-id, sku, supplier invoice", "Evidence packet for inbound, valuation, and disputed claim cases.", "Store as PDF/JPG links, not inside n8n credentials."],
]


claim_rules = [
    ["INBOUND_SHORTAGE", "Shipment to Amazon claims", "Shipped qty is greater than Amazon received/reconciled qty after claim window opens.", "Inbound shipments, ledger, reimbursements, invoices, POD/BOL", "Often 15-60 days from delivery; verify marketplace policy.", "Shipment ID, SKU/FNSKU, expected qty, received qty, invoice, packing slip, POD/BOL.", "Create case packet when evidence complete.", "Yes", "Do not file before shipment is eligible for investigation.", "Case ID, deadline, missing qty, expected recovery, evidence links."],
    ["FC_LOST_UNRECONCILED", "Fulfillment center operations", "Ledger has lost/misplaced/unreconciled quantity with no matching found or reimbursement.", "Inventory ledger, reimbursements, manage inventory", "No later than 60 days after item reported lost/damaged.", "Ledger rows, ReferenceID, FNSKU, quantity, no reimbursement proof.", "Rank by deadline and recovery.", "Yes", "Amazon says many warehouse lost/damaged cases are proactive, so dedupe first.", "ReferenceID, unreconciled qty, reason code, reimbursement status."],
    ["FC_DAMAGED_UNREIMBURSED", "Fulfillment center operations", "Ledger damage event caused under Amazon control, no reimbursement/replacement.", "Inventory ledger, reimbursements", "No later than 60 days after reported damaged.", "Reason code, disposition, ReferenceID, FNSKU, quantity.", "Create claim only when not already reconciled.", "Yes", "Exclude customer-damaged/defective seller-fault cases.", "Damage type, unit status, evidence checklist."],
    ["CUSTOMER_REFUND_NO_RETURN", "Customer return claims", "Refund/replacement issued, item not returned to FC and no reimbursement after waiting period.", "Finances, FBA returns, shipments, reimbursements", "60-120 days after customer refund/replacement.", "Order ID, refund date, refund amount, no return row, no reimbursement row.", "Create case packet once day 60 passes.", "Yes", "Do not file before day 60.", "Order ID, refund event, deadline, recovery estimate."],
    ["CUSTOMER_RETURN_UNSELLABLE_AMAZON_FAULT", "Customer return claims", "Returned item unsellable due carrier/Amazon handling, not customer damage, and no reimbursement.", "FBA returns, inventory ledger, finances, reimbursements", "60-120 days after refund/replacement.", "Order ID, LPN, return reason, disposition, photos if removed, ledger rows.", "Flag for review; evidence sensitivity high.", "Yes", "Customer-damaged, defective, recall, or policy-violating items are usually excluded.", "LPN, disposition, reason, reviewer decision."],
    ["REPLACEMENT_NO_RETURN", "Customer return claims", "Replacement issued; original order item not returned or reimbursed.", "Replacements, returns, finances, reimbursements", "60-120 days after replacement/refund.", "Original order ID, replacement order ID, dates, return/reimbursement checks.", "Create case packet after waiting period.", "Yes", "Check if replacement reason is excluded.", "Original order, replacement order, replacement reason."],
    ["REMOVAL_LOST_IN_TRANSIT", "Removals claims", "Removal shipment created/shipped but tracking not delivered and no reimbursement.", "Removal order detail, removal shipment detail, reimbursements", "15-75 days from removal shipment creation.", "Removal order ID, tracking, shipped qty, carrier, no delivery proof.", "Escalate quickly due short window.", "Yes", "Do not file before day 15.", "Removal ID, tracking, carrier, deadline."],
    ["REMOVAL_DAMAGED_DIFFERENT_INCOMPLETE", "Removals claims", "Removal delivered but units arrived damaged, wrong, or incomplete.", "Removal reports, seller receiving log, photos", "Within 60 days after delivery; some older guidance cites 30 days for certain removal receipt issues.", "Photos, packing material, received qty, SKU/FNSKU, delivery date.", "Generate evidence checklist and draft.", "Yes", "Requires seller receiving inspection process.", "Photo links, received condition, quantity variance."],
    ["REIMBURSEMENT_APPROVED_NOT_PAID", "Payment follow-up", "Case or reimbursement tool says approved, but no reimbursement appears in reimbursement/settlement/financial events.", "Cases, reimbursements, finances, settlement", "Follow up after 4-5 business days; verify current policy.", "Case ID, approval message, reimbursement report check, settlement check.", "Create follow-up task, not new claim.", "Yes", "Avoid duplicate reimbursement claim; this is payment reconciliation.", "Case ID, missing payment proof, follow-up age."],
    ["REIMBURSEMENT_UNDERPAID", "Valuation dispute", "Reimbursement amount is below submitted/approved manufacturing cost or expected post-order value.", "Reimbursements, COGS, invoices, finances", "Often 60 days after reimbursement for US valuation dispute; verify marketplace.", "Reimbursement ID, amount, expected value calculation, proof-of-value docs.", "Create valuation dispute packet.", "Yes", "Do not inflate manufacturing cost; Amazon can reject outliers.", "Expected amount, actual amount, variance, proof docs."],
    ["REIMBURSEMENT_REVERSED", "Payment follow-up", "Amazon reversed a reimbursement; item was not returned to inventory or reversal seems unmatched.", "Reimbursements, ledger, finances, settlement", "Review immediately.", "Original reimbursement ID, retraction event, ledger found/reversal rows.", "Create exception review.", "Yes", "Some reversals are valid when inventory is found.", "Original ID, reversal ID, net impact."],
    ["FBA_FEE_OVERCHARGE", "Fee issue", "Charged fulfillment fee exceeds expected fee tier based on dimensions/weight.", "Fee preview, settlement/finances, catalog dimensions", "Commonly short windows; verify current policy before filing.", "SKU, ASIN, charged fee, expected fee, dimensions, weight evidence.", "Aggregate by SKU and period.", "Yes", "Requires reliable dimension/weight baseline.", "Variance, affected orders, fee basis."],
    ["STORAGE_FEE_OVERCHARGE", "Fee issue", "Storage or aged inventory fee inconsistent with inventory qty, age, volume, or fee rate.", "Storage fee reports, inventory planning, manage inventory", "Verify current fee dispute policy.", "ASIN/FNSKU, month, qty charged, volume, rate, expected amount.", "Create monthly fee review.", "Yes", "Usually lower confidence than inventory loss cases.", "Month, ASIN, variance, evidence."],
    ["REFUND_FEE_MISMATCH", "Payment issue", "Refund, commission, or FBA fee reversal does not reconcile with expected transaction economics.", "Finances, settlement, shipments, returns", "Verify current policy and claim path.", "Order ID, refund event, original sale event, fee comparison.", "Flag for accounting review.", "Yes", "Complex and easy to false-positive; aggregate before action.", "Order economics, expected vs actual."],
    ["STRANDED_OR_DEFECTIVE_MISCLASSIFIED", "Inventory support", "Inventory marked defective/stranded in a way that blocks sale or reimbursement eligibility unexpectedly.", "Stranded inventory, manage inventory, ledger", "Operational issue; claim window depends on cause.", "SKU/FNSKU, stranded reason, listing status, inventory history.", "Create support packet, not reimbursement claim first.", "Yes", "May not be a reimbursement claim until loss/damage is proven.", "Root cause, support path, action owner."],
]


audit_rows = [
    ["CASE-0001", "Ready", "High", "CUSTOMER_REFUND_NO_RETURN", "US", date(2026, 5, 28), None, 124.80, 0.92, "CAM-KIT-01", "X001CAM01", "B0EXAMPLE1", "113-0000001-0000001", "Complete", "File customer return claim", "FBA customer return not received: order 113-0000001-0000001", "Refund posted 2026-03-28; no FBA return row; no reimbursement row."],
    ["CASE-0002", "Ready", "High", "FC_LOST_UNRECONCILED", "US", date(2026, 5, 23), None, 356.40, 0.95, "HEADSET-BLK", "X001HEADB", "B0EXAMPLE2", "ADJ-784512", "Complete", "File FC operations claim", "Unreconciled misplaced FBA inventory: X001HEADB", "Ledger reason M, unreconciled qty 3, no found/reimbursement within review period."],
    ["CASE-0003", "Review", "Medium", "REIMBURSEMENT_UNDERPAID", "US", date(2026, 6, 2), None, 71.30, 0.74, "FILTER-2PK", "X001FILTR", "B0EXAMPLE3", "RMB-3355901", "Needs invoice", "Reviewer validate COGS", "Reimbursement valuation dispute: RMB-3355901", "Amazon estimate below seller manufacturing cost; invoice not attached yet."],
    ["CASE-0004", "Hold", "Medium", "INBOUND_SHORTAGE", "US", date(2026, 6, 11), None, 512.00, 0.81, "LAMP-WHT", "X001LAMPW", "B0EXAMPLE4", "FBA18ABCD1", "Needs POD/BOL", "Wait for evidence upload", "Shipment to Amazon missing units: FBA18ABCD1", "Sent 80, received 64, shipment eligible; proof of delivery missing."],
    ["CASE-0005", "Ready", "High", "REMOVAL_LOST_IN_TRANSIT", "US", date(2026, 5, 21), None, 198.95, 0.88, "MIC-USB", "X001MICUS", "B0EXAMPLE5", "RMV-77811", "Complete", "File removal lost in transit claim", "Removal shipment lost in transit: RMV-77811", "Carrier tracking has no delivery scan; day 15 passed and day 75 approaching."],
    ["CASE-0006", "Closed - already reimbursed", "Low", "FC_DAMAGED_UNREIMBURSED", "US", date(2026, 5, 30), None, 0.00, 0.99, "BAG-GRN", "X001BAGGN", "B0EXAMPLE6", "ADJ-910012", "Complete", "No action", "No claim needed: already reimbursed", "Reimbursement report contains matching reimbursement ID."],
    ["CASE-0007", "Review", "High", "FBA_FEE_OVERCHARGE", "US", date(2026, 6, 5), None, 89.60, 0.67, "STAND-ALU", "X001STAND", "B0EXAMPLE7", "FEE-APR-2026", "Needs dimensions", "Validate dimension proof", "FBA fulfillment fee overcharge review: STAND-ALU", "Charged large standard; catalog dimensions suggest small standard."],
    ["CASE-0008", "Draft", "Medium", "CUSTOMER_RETURN_UNSELLABLE_AMAZON_FAULT", "US", date(2026, 6, 9), None, 249.99, 0.76, "DRONE-MINI", "X001DRONE", "B0EXAMPLE8", "112-0000008-0000008", "Needs photos", "Request removal/photos then review", "Unsellable FBA return under Amazon/carrier control", "Return reason indicates carrier damage; disposition unsellable; photos pending."],
    ["CASE-0009", "Investigate", "Medium", "REPLACEMENT_NO_RETURN", "US", date(2026, 6, 15), None, 139.45, 0.70, "SCALE-DIG", "X001SCALE", "B0EXAMPLE9", "113-0000009-0000009", "Partial", "Cross-check replacement reason", "Replacement order original not returned: 113-0000009-0000009", "Replacement issued; original return not found in returns report."],
    ["CASE-0010", "Ready", "Low", "STORAGE_FEE_OVERCHARGE", "US", date(2026, 6, 30), None, 43.20, 0.61, "CASE-PACK", "X001CASEP", "B0EXAMP10", "STO-2026-04", "Complete", "Batch into monthly fee case", "Monthly storage fee variance: April 2026", "Small variance; hold until monthly batch exceeds threshold."],
]


evidence_rows = [
    ["Inbound shortage", "Shipment ID, SKU/FNSKU, ASIN", "Supplier invoice, packing slip, POD/BOL, box content docs", "Shipment eligible, not deleted/cancelled, qty mismatch remains, not reimbursed.", "Shipment, expected qty, received qty, missing qty, proof summary.", "PDF invoice, POD/BOL, packing slip.", "Shipment still receiving or Amazon says wait; missing docs.", "Human reviewer initials and date."],
    ["FC lost/misplaced", "ReferenceID, FNSKU, SKU, ASIN", "Ledger extract, reimbursement report extract", "Unreconciled qty remains; no found/reimbursement row.", "Date, reason code, qty, no matching reimbursement.", "Ledger CSV/PDF, reimbursement check screenshot if manual.", "Found row or reimbursement already exists.", "Human reviewer initials and date."],
    ["FC damaged", "ReferenceID, FNSKU, disposition", "Ledger extract, reimbursement report extract", "Amazon/facility/carrier damage, not customer damage or defective.", "Damage event, quantity, status, claim window.", "Ledger CSV/PDF.", "Customer damage, recall, seller defect, policy violation.", "Human reviewer initials and date."],
    ["Customer refund no return", "Order ID, SKU/FNSKU, refund event", "Finances refund row, returns report, reimbursement report", "Day 60 passed; day 120 not passed; no return/reimbursement.", "Refund date, amount, missing return evidence.", "Finances/returns/reimbursements extracts.", "Before day 60 or non-returnable policy exclusion.", "Human reviewer initials and date."],
    ["Unsellable customer return", "Order ID, LPN, return date", "Return report, photos/removal inspection, ledger rows", "Reason/disposition supports Amazon/carrier responsibility.", "Condition timeline and responsibility argument.", "Return report, photos, LPN evidence.", "Customer damaged, defective, missing seal/packaging exclusions.", "Human reviewer initials and date."],
    ["Removal lost/damaged", "Removal order ID, tracking, SKU/FNSKU", "Removal reports, delivery proof or no-delivery tracking, photos", "Inside removal claim window and no reimbursement.", "Removal shipment timeline and affected quantity.", "Tracking PDF, photos, receiving log.", "Carrier delivered correctly and seller lacks inspection proof.", "Human reviewer initials and date."],
    ["Fee overcharge", "Order ID or monthly fee ID, SKU/ASIN", "Fee report, settlement row, dimensions/weight proof", "Expected fee calculation is documented and material.", "Charged fee, expected fee, variance, proof basis.", "Dimension proof, rate card references, fee rows.", "Dimension data uncertain or variance immaterial.", "Human reviewer initials and date."],
]


n8n_rows = [
    ["1. Schedule", "Cron", "Marketplace list and cadence", "Run ID", "Skip if previous run active", "Automation", "Daily 04:00", "Keep daily cadence because FBA reports update daily or near real-time."],
    ["2. Pull Reports", "HTTP Request", "SP-API credentials, report types, date ranges", "Report document IDs", "Retry 429 with backoff", "Automation", "Daily/monthly", "Request only needed windows to avoid report throttling."],
    ["3. Normalize", "Code", "Tab-delimited reports", "Standardized tables", "Validate headers and required fields", "Automation", "Every run", "Map all dates to UTC and store raw + normalized rows."],
    ["4. Store Raw Data", "Postgres / Sheet / Drive", "Raw report files", "Audit trail", "Checksum duplicate files", "Automation", "Every run", "Keep source files for defensible evidence."],
    ["5. Rule Engine", "Code", "Normalized tables + policy windows", "Candidate cases", "Reject if missing keys", "Automation", "Every run", "Rules should be deterministic before any AI drafting."],
    ["6. Deduplicate", "Database Query", "Candidate cases, reimbursement rows, case tracker", "Filtered queue", "Block duplicate order/reference IDs", "Automation", "Every run", "This is the most important safety step."],
    ["7. Evidence Builder", "Code", "Candidate case + source rows", "Case packet", "Mark Needs Evidence if docs absent", "Automation", "Every run", "No case packet should be marked Ready without evidence."],
    ["8. AI Draft", "OpenAI / LLM Node", "Case packet only", "Subject and body draft", "Use strict template and no unsupported claims", "Automation", "Optional", "Draft concise factual case text; no fabricated screenshots or documents."],
    ["9. Human Review", "Slack / Email / Sheet", "Ready cases", "Approved / rejected / hold", "Require reviewer signoff", "Operations", "Daily", "Human approves before Seller Central case creation."],
    ["10. Case Filing", "Manual or browser-assisted", "Approved case packet", "Seller Central case ID", "Capture case ID immediately", "Operations", "On approval", "Keep manual first; later consider guarded automation."],
    ["11. Follow-up", "Cron + SP-API Pull", "Open case tracker", "Recovered / denied / pending status", "Escalate silent approvals", "Automation", "Daily", "Match to reimbursements and financial events."],
    ["12. Close Loop", "Database Update", "Reimbursement/payment confirmation", "Final result and ROI", "Flag reversals", "Automation", "Every run", "Track recovered amount, denied amount, and reversals."],
]


template_rows = [
    ["CUSTOMER_REFUND_NO_RETURN", "FBA customer return not received: order {order_id}", "I am requesting review of an FBA customer return where Amazon refunded or replaced the customer, but the item has not been returned to inventory and no reimbursement is present.", "Refund date; Order ID; SKU/FNSKU; return report check; reimbursement report check.", "Please reimburse the eligible amount or confirm the specific policy reason this order is not eligible."],
    ["FC_LOST_UNRECONCILED", "Unreconciled misplaced FBA inventory: {fnsku}", "I am requesting reimbursement review for FBA inventory marked lost or misplaced at the fulfillment center with unreconciled quantity still outstanding.", "Ledger date; ReferenceID; Reason; FNSKU; quantity; no found/reimbursement match.", "Please reimburse or reconcile the missing units."],
    ["INBOUND_SHORTAGE", "Shipment to Amazon missing units: {shipment_id}", "I am requesting investigation of units shipped to Amazon but not received/reconciled for this FBA shipment.", "Shipment ID; expected qty; received qty; SKU/FNSKU; invoice; packing slip; proof of delivery.", "Please locate, receive, or reimburse the eligible missing units."],
    ["REMOVAL_LOST_IN_TRANSIT", "Removal shipment lost in transit: {removal_order_id}", "I am requesting review of a removal shipment that was shipped but not delivered back to me and has not been reimbursed.", "Removal order ID; tracking; ship date; SKU/FNSKU; quantity; no delivery scan.", "Please reimburse the eligible lost removal shipment units."],
    ["REIMBURSEMENT_UNDERPAID", "Reimbursement valuation dispute: {reimbursement_id}", "I am requesting review of the reimbursement valuation because the issued amount appears lower than the documented eligible value.", "Reimbursement ID; actual amount; expected amount; manufacturing cost evidence; invoice.", "Please adjust the reimbursement or provide the calculation basis."],
]


reason_code_rows = [
    ["M", "Inventory misplaced", "Decrease", "Potential FC lost/misplaced claim if unreconciled and not reimbursed.", "Match against F/N found events and reimbursement rows before filing."],
    ["F", "Inventory found", "Increase", "Usually reconciles a prior lost/misplaced event.", "If found after reimbursement, Amazon may reverse the reimbursement."],
    ["N", "Inventory found / correction", "Increase", "May indicate inventory transferred to account or reimbursement-related correction.", "Review with surrounding ledger rows."],
    ["D", "Inventory disposed of", "Decrease", "Potential removal/disposal issue if not seller-requested or policy-driven.", "Check disposal reason and reimbursement status."],
    ["E", "Damaged at Amazon fulfillment center", "Decrease", "Potential FC damaged claim if no reimbursement/replacement.", "Commonly followed by P disposition change."],
    ["6 / 7", "Damaged at Amazon fulfillment center", "Decrease", "Damage movement from carrier/expired disposition into FC-damaged state.", "Often followed by P; verify disposition and responsibility."],
    ["H / K / U", "Damaged at Amazon fulfillment center", "Decrease", "Damage movement from customer-damaged, defective, or distributor-damaged state.", "High review risk; customer/seller fault may be excluded."],
    ["P", "Inventory disposition change", "Increase", "Usually paired with damage/disposition decrease codes.", "P by itself is not a claim; inspect the preceding decrease event."],
    ["Q", "Inventory disposition change", "Decrease", "Disposition removed before being added back as another disposition.", "Usually followed by P. This is probably one of the letters you remembered."],
    ["O", "Inventory correction", "Decrease", "May reflect inventory transferred out or reimbursement-related correction.", "Check if this closes a reimbursement or creates a new discrepancy."],
    ["3 / 4", "Product redefinition transfer in/out", "Increase / Decrease", "SKU identity or mapping correction, not automatically reimbursable.", "Useful for explaining why inventory moved between identifiers."],
]


sources = [
    ["Amazon SP-API FBA Reports", "https://developer-docs.amazon.com/sp-api/lang-en_EN/docs/report-type-values-fba", "Report types, fields, cadence notes.", "Official developer documentation."],
    ["Amazon Settlement Reports", "https://developer-docs.amazon.com/sp-api/lang-en_EN/docs/report-type-values-settlement", "Settlement report V2 fields.", "Official developer documentation."],
    ["Amazon Finances API listFinancialEvents", "https://developer-docs.amazon.com/sp-api/reference/listfinancialevents", "Financial event retrieval, delays, rate limits.", "Official developer documentation."],
    ["Amazon Returns Reports", "https://developer-docs.amazon.com/sp-api/docs/report-type-values-returns", "Return reports for FBM/hybrid extension.", "Official developer documentation."],
    ["Amazon FBA Inventory Reimbursement Policy PDF US", "https://m.media-amazon.com/images/G/01/rainier/help./PC__Kiwi_Comms_redlined_policy_US_clean.pdf", "Eligibility, reimbursement value, manufacturing cost, cautions.", "Official Amazon PDF."],
    ["Amazon Reimbursement Window Update", "https://sellercentral.amazon.com/seller-forums/discussions/t/81c3235d-4c44-47ba-96c5-883cecab3244", "60-day FC operations, 60-120 day customer return, removals windows.", "Official Seller Forums announcement."],
    ["Inventory Ledger Reason Code Reference", "https://help.reasonautomation.com/seller/inventory-ledger-details", "Practical mapping of M, F, N, D, E, P, Q, and related reason codes.", "Third-party data dictionary; verify against live Amazon report labels."],
    ["GETIDA", "https://getida.com/", "Market reference for reimbursement service categories.", "Third-party service reference."],
    ["SellerVault FBA Reimbursement Recovery", "https://sellervault.io/fba-reimbursement-recovery", "Market reference for detection algorithm categories.", "Third-party service reference."],
    ["ReimburseOps", "https://reimburseops.com/", "Market reference for upload/audit style reimbursement tooling.", "Third-party service reference."],
]


def style_range(ws, cell_range, fill=None, font=None, border=None, alignment=None):
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment


def add_section_title(ws, row, title, end_col=17):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
    cell.font = Font(color=COLORS["white"], bold=True, size=13)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24


def add_table(ws, name, start_row, start_col, headers, rows, style_name="TableStyleMedium2"):
    for col, header in enumerate(headers, start_col):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor=COLORS["blue"])
        cell.font = Font(color=COLORS["white"], bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row_data in enumerate(rows, start_row + 1):
        for c_idx, value in enumerate(row_data, start_col):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(value, float):
                cell.number_format = "0.00"
            if isinstance(value, date):
                cell.number_format = "yyyy-mm-dd"
    end_row = start_row + len(rows)
    end_col = start_col + len(headers) - 1
    ref = f"{ws.cell(start_row, start_col).coordinate}:{ws.cell(end_row, end_col).coordinate}"
    tab = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name=style_name, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    return end_row


def apply_borders(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color=COLORS["border"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "FBA Blueprint"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    widths = {
        "A": 20, "B": 24, "C": 30, "D": 28, "E": 22, "F": 24, "G": 18, "H": 18,
        "I": 17, "J": 24, "K": 18, "L": 18, "M": 28, "N": 25, "O": 32, "P": 42, "Q": 48,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("A1:Q1")
    ws["A1"] = "Amazon FBA Reimbursement n8n Workflow - Excel Planning & Report Example"
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["navy"])
    ws["A1"].font = Font(color=COLORS["white"], bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:Q2")
    ws["A2"] = "Planning-only template. The workflow should identify evidence-backed reimbursement candidates, dedupe against existing reimbursements/cases, and route drafts for human approval before filing."
    ws["A2"].fill = PatternFill("solid", fgColor=COLORS["gray"])
    ws["A2"].font = Font(color=COLORS["dark_text"], italic=True)
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 34

    ws["A3"] = "As of date"
    ws["B3"] = AS_OF
    ws["B3"].number_format = "yyyy-mm-dd"
    ws["D3"] = "Operating principle"
    ws["E3"] = "Generate case packets, not blind automated case spam."
    ws["E3"].font = Font(bold=True, color=COLORS["red"])
    style_range(ws, "A3:Q3", fill=PatternFill("solid", fgColor="EAF0F6"), alignment=Alignment(vertical="center"))

    kpis = [
        ("A5:C5", "A6:C6", "Case types covered", "=COUNTA(A33:A47)"),
        ("D5:F5", "D6:F6", "Source reports/APIs", "=COUNTA(C10:C25)"),
        ("G5:I5", "G6:I6", "Sample open recovery", '=SUMIF(B54:B63,"<>Closed - already reimbursed",H54:H63)'),
        ("J5:L5", "J6:L6", "Due in 14 days", '=COUNTIFS(G54:G63,"<=14",B54:B63,"<>Closed - already reimbursed")'),
        ("M5:Q5", "M6:Q6", "Automation mode", "Human-approved case packets"),
    ]
    for label_range, value_range, label, value in kpis:
        ws.merge_cells(label_range)
        ws.merge_cells(value_range)
        label_cell = ws[label_range.split(":")[0]]
        value_cell = ws[value_range.split(":")[0]]
        label_cell.value = label
        value_cell.value = value
        for merged_range in [label_range, value_range]:
            top_left = merged_range.split(":")[0]
            cell = ws[top_left]
            cell.fill = PatternFill("solid", fgColor=COLORS["light_blue"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, color=COLORS["navy"])
        value_cell.font = Font(bold=True, size=14, color=COLORS["navy"])
        if "recovery" in label.lower():
            value_cell.number_format = "$#,##0.00"

    add_section_title(ws, 8, "1. Source Reports to Download", end_col=8)
    source_headers = ["Priority", "System", "Report/API", "Report Type / Endpoint", "Cadence", "Key Join Fields", "Used For / Why", "Notes"]
    source_end = add_table(ws, "tblSourceReports", 9, 1, source_headers, source_reports, "TableStyleMedium2")
    apply_borders(ws, 9, source_end, 1, 8)

    ws.merge_cells("J8:L8")
    ws["J8"] = "Summary by Topic"
    ws["J8"].fill = PatternFill("solid", fgColor=COLORS["teal"])
    ws["J8"].font = Font(color=COLORS["white"], bold=True)
    ws["J8"].alignment = Alignment(horizontal="center")
    summary_headers = ["Topic Code", "Count", "Est Recovery"]
    for idx, header in enumerate(summary_headers, 10):
        ws.cell(9, idx, header)
        ws.cell(9, idx).fill = PatternFill("solid", fgColor=COLORS["teal"])
        ws.cell(9, idx).font = Font(color=COLORS["white"], bold=True)
        ws.cell(9, idx).alignment = Alignment(horizontal="center")
    summary_topics = [
        "CUSTOMER_REFUND_NO_RETURN", "FC_LOST_UNRECONCILED", "INBOUND_SHORTAGE",
        "REMOVAL_LOST_IN_TRANSIT", "REIMBURSEMENT_UNDERPAID", "FBA_FEE_OVERCHARGE",
        "CUSTOMER_RETURN_UNSELLABLE_AMAZON_FAULT", "REPLACEMENT_NO_RETURN", "STORAGE_FEE_OVERCHARGE",
    ]
    for row_idx, topic in enumerate(summary_topics, 10):
        ws.cell(row_idx, 10, topic)
        ws.cell(row_idx, 11, f'=COUNTIF($D$54:$D$63,J{row_idx})')
        ws.cell(row_idx, 12, f'=SUMIF($D$54:$D$63,J{row_idx},$H$54:$H$63)')
        ws.cell(row_idx, 12).number_format = "$#,##0.00"
        for col in range(10, 13):
            ws.cell(row_idx, col).alignment = Alignment(wrap_text=True)
    apply_borders(ws, 9, 18, 10, 12)

    chart = BarChart()
    chart.title = "Estimated Recovery by Topic"
    chart.y_axis.title = "USD"
    chart.x_axis.title = "Topic"
    chart.height = 8
    chart.width = 14
    data = Reference(ws, min_col=12, min_row=9, max_row=18)
    cats = Reference(ws, min_col=10, min_row=10, max_row=18)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.style = 10
    ws.add_chart(chart, "N8")

    add_section_title(ws, 31, "2. Discrepancy Rules / Case Topics")
    rule_headers = ["Topic Code", "Amazon Case Path", "Trigger Logic", "Source Reports", "Claim Window", "Evidence Required", "Auto Action", "Human Approval", "Risk Notes", "Output Fields"]
    rules_end = add_table(ws, "tblClaimRules", 32, 1, rule_headers, claim_rules, "TableStyleMedium4")
    apply_borders(ws, 32, rules_end, 1, 10)

    add_section_title(ws, 52, "3. Example Audit Queue - What the Final Report Looks Like")
    audit_headers = ["Case ID", "Status", "Priority", "Topic Code", "Marketplace", "Deadline", "Days Left", "Est Recovery", "Confidence", "SKU", "FNSKU", "ASIN", "Order/Shipment/Removal ID", "Evidence Status", "Next Action", "Draft Subject", "Notes"]
    audit_start = 53
    for col, header in enumerate(audit_headers, 1):
        cell = ws.cell(audit_start, col, header)
        cell.fill = PatternFill("solid", fgColor=COLORS["red"])
        cell.font = Font(color=COLORS["white"], bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row_data in enumerate(audit_rows, audit_start + 1):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c_idx == 7:
                cell.value = f"=F{r_idx}-$B$3"
                cell.number_format = "0"
            elif c_idx == 8:
                cell.number_format = "$#,##0.00"
            elif c_idx == 9:
                cell.number_format = "0%"
            elif isinstance(value, date):
                cell.number_format = "yyyy-mm-dd"
    audit_end = audit_start + len(audit_rows)
    tab = Table(displayName="tblAuditQueue", ref=f"A{audit_start}:Q{audit_end}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    apply_borders(ws, audit_start, audit_end, 1, 17)

    status_dv = DataValidation(type="list", formula1='"Ready,Review,Draft,Hold,Investigate,Closed - already reimbursed"', allow_blank=False)
    priority_dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=False)
    evidence_dv = DataValidation(type="list", formula1='"Complete,Partial,Needs invoice,Needs POD/BOL,Needs photos,Needs dimensions"', allow_blank=False)
    ws.add_data_validation(status_dv)
    ws.add_data_validation(priority_dv)
    ws.add_data_validation(evidence_dv)
    status_dv.add(f"B54:B{audit_end}")
    priority_dv.add(f"C54:C{audit_end}")
    evidence_dv.add(f"N54:N{audit_end}")

    ws.conditional_formatting.add(f"C54:C{audit_end}", FormulaRule(formula=['C54="High"'], fill=PatternFill("solid", fgColor=COLORS["light_red"])))
    ws.conditional_formatting.add(f"C54:C{audit_end}", FormulaRule(formula=['C54="Medium"'], fill=PatternFill("solid", fgColor=COLORS["yellow"])))
    ws.conditional_formatting.add(f"C54:C{audit_end}", FormulaRule(formula=['C54="Low"'], fill=PatternFill("solid", fgColor=COLORS["light_green"])))
    ws.conditional_formatting.add(f"G54:G{audit_end}", FormulaRule(formula=["AND($G54<=7,$B54<>\"Closed - already reimbursed\")"], fill=PatternFill("solid", fgColor=COLORS["light_red"])))
    ws.conditional_formatting.add(f"G54:G{audit_end}", FormulaRule(formula=["AND($G54>7,$G54<=14,$B54<>\"Closed - already reimbursed\")"], fill=PatternFill("solid", fgColor=COLORS["yellow"])))
    ws.conditional_formatting.add(f"B54:B{audit_end}", FormulaRule(formula=['B54="Ready"'], fill=PatternFill("solid", fgColor=COLORS["light_green"])))
    ws.conditional_formatting.add(f"B54:B{audit_end}", FormulaRule(formula=['B54="Closed - already reimbursed"'], fill=PatternFill("solid", fgColor="D9D9D9")))

    next_row = audit_end + 3
    add_section_title(ws, next_row, "4. Evidence Checklist / Case Packet Template")
    evidence_headers = ["Claim Family", "Required IDs", "Required Docs", "Validation Before Case", "Case Text Fields", "Attachments", "Do Not File When", "Reviewer Signoff"]
    evidence_end = add_table(ws, "tblEvidenceChecklist", next_row + 1, 1, evidence_headers, evidence_rows, "TableStyleMedium7")
    apply_borders(ws, next_row + 1, evidence_end, 1, 8)

    next_row = evidence_end + 3
    add_section_title(ws, next_row, "5. n8n Workflow Map")
    n8n_headers = ["Stage", "n8n Node Type", "Inputs", "Outputs", "Error Handling", "Owner", "Frequency", "Notes"]
    n8n_end = add_table(ws, "tblN8nMap", next_row + 1, 1, n8n_headers, n8n_rows, "TableStyleMedium6")
    apply_borders(ws, next_row + 1, n8n_end, 1, 8)

    next_row = n8n_end + 3
    add_section_title(ws, next_row, "6. Case Draft Templates")
    template_headers = ["Topic Code", "Subject Pattern", "Opening Sentence", "Evidence Bullets", "Requested Outcome"]
    template_end = add_table(ws, "tblDraftTemplates", next_row + 1, 1, template_headers, template_rows, "TableStyleMedium3")
    apply_borders(ws, next_row + 1, template_end, 1, 5)

    next_row = template_end + 3
    add_section_title(ws, next_row, "7. Inventory Reason Code Legend")
    reason_headers = ["Code", "Meaning", "Movement", "Reimbursement Interpretation", "Notes"]
    reason_end = add_table(ws, "tblReasonCodes", next_row + 1, 1, reason_headers, reason_code_rows, "TableStyleMedium8")
    apply_borders(ws, next_row + 1, reason_end, 1, 5)

    next_row = reason_end + 3
    add_section_title(ws, next_row, "8. Sources & Policy References")
    source_ref_headers = ["Source", "URL", "Used For", "Notes"]
    source_ref_end = add_table(ws, "tblSources", next_row + 1, 1, source_ref_headers, sources, "TableStyleMedium5")
    apply_borders(ws, next_row + 1, source_ref_end, 1, 4)

    for row in range(1, source_ref_end + 1):
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 16, 18)
    for row in range(9, source_end + 1):
        ws.row_dimensions[row].height = 52
    for row in range(32, rules_end + 1):
        ws.row_dimensions[row].height = 64
    for row in range(54, audit_end + 1):
        ws.row_dimensions[row].height = 46
    for row in range(audit_end + 4, source_ref_end + 1):
        if ws.row_dimensions[row].height < 44:
            ws.row_dimensions[row].height = 44

    # Number formatting for summary cards and body columns.
    for cell in ["G6", "L10", "L11", "L12", "L13", "L14", "L15", "L16", "L17", "L18"]:
        ws[cell].number_format = "$#,##0.00"

    # Gentle body font.
    for row in ws.iter_rows(min_row=1, max_row=source_ref_end, min_col=1, max_col=17):
        for cell in row:
            if cell.font == Font():
                cell.font = Font(color=COLORS["dark_text"], size=10)
            cell.alignment = cell.alignment.copy(vertical=cell.alignment.vertical or "top")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(OUTPUT.resolve())


if __name__ == "__main__":
    build()
