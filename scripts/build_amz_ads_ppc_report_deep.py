import math
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.append(str(Path(__file__).parent))
import build_amz_ads_ppc_report as base


DATA_DIR = Path("/Users/v/Documents/ART_GPT/outputs/amazon_ads_ppc_20260607")
OUT_FILE = DATA_DIR / "amazon_ads_ppc_keyword_report_DEEP_2026-06-07.xlsx"


def norm_text(value):
    return re.sub(r"\s+", " ", str(value or "").strip().lower().replace('"', ""))


def extract_asin(value):
    s = str(value or "").upper()
    m = re.search(r'\b(B[0-9A-Z]{9})\b', s)
    return m.group(1) if m else ""


def num(df, cols):
    for col in cols:
        if col not in df:
            df[col] = 0
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def enrich_extra(df):
    if df is None or len(df) == 0:
        return pd.DataFrame()
    df = df.copy()
    df = num(df, [
        "impressions", "clicks", "cost", "spend", "sales7d", "sales14d",
        "purchases7d", "purchases14d", "unitsSoldClicks7d", "unitsSoldClicks14d",
        "attributedSalesSameSku7d", "salesOtherSku7d", "purchasesSameSku7d",
        "purchasesOtherSku7d", "unitsSoldSameSku7d", "unitsSoldOtherSku7d",
        "costPerClick", "clickThroughRate", "topOfSearchImpressionShare",
    ])
    df["ctr_calc"] = df["clicks"] / df["impressions"].replace(0, pd.NA)
    df["cpc_calc"] = df["cost"] / df["clicks"].replace(0, pd.NA)
    df["cvr_calc"] = df["purchases7d"] / df["clicks"].replace(0, pd.NA)
    df["acos_calc"] = df["cost"] / df["sales7d"].replace(0, pd.NA)
    df["roas_calc"] = df["sales7d"] / df["cost"].replace(0, pd.NA)
    units = df["unitsSoldClicks7d"].where(df["unitsSoldClicks7d"] > 0, df["purchases7d"])
    df["projected_pre_ad_profit"] = df["sales7d"] * base.NET_AFTER_AMAZON_RATE - units * base.UNIT_COST
    df["projected_after_ad_profit"] = df["projected_pre_ad_profit"] - df["cost"]
    return df


def load_extra(name):
    return enrich_extra(base.load_json(name))


def write_df(ws, df, freeze=True, max_rows=None, start_row=1, start_col=1):
    if df is None:
        df = pd.DataFrame()
    use = df.head(max_rows).copy() if max_rows else df.copy()
    use = use.where(pd.notna(use), "")
    if use.empty:
        ws.cell(start_row, start_col, "No data returned").font = Font(bold=True)
        return
    for j, col in enumerate(use.columns, start_col):
        cell = ws.cell(start_row, j, str(col))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for i, row in enumerate(use.itertuples(index=False), start_row + 1):
        for j, value in enumerate(row, start_col):
            if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
                value = ""
            ws.cell(i, j, value)
    ws.freeze_panes = ws.cell(start_row + 1, start_col).coordinate if freeze else None
    ws.auto_filter.ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(start_col + use.shape[1] - 1)}{start_row + use.shape[0]}"
    style_sheet(ws)
    style_numbers(ws)


def style_sheet(ws):
    thin = Side(style="thin", color="E5E7EB")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for col in range(1, min(ws.max_column, 35) + 1):
        max_len = 8
        for row in range(1, min(ws.max_row, 600) + 1):
            val = ws.cell(row, col).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 55))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 10), 44)
    ws.sheet_view.showGridLines = False


def style_numbers(ws):
    money_keywords = ("cost", "spend", "sales", "profit", "cpc", "bid", "budget", "revenue", "amount")
    pct_keywords = ("acos", "ctr", "cvr", "roas", "share", "rate")
    headers = {str(ws.cell(1, c).value or "").lower(): c for c in range(1, ws.max_column + 1)}
    for name, col in headers.items():
        fmt = None
        if any(k in name for k in money_keywords) and "roas" not in name:
            fmt = '$#,##0.00'
        if any(k in name for k in pct_keywords) and "roas" not in name:
            fmt = "0.0%"
        if "roas" in name:
            fmt = "0.00"
        if fmt:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).number_format = fmt


def add_title(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="111827")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].alignment = Alignment(wrap_text=True)


def totals_row(label, df):
    cols = ["impressions", "clicks", "cost", "spend", "sales7d", "purchases7d", "unitsSoldClicks7d"]
    row = {"source": label, "rows": len(df)}
    for c in cols:
        row[c] = float(df[c].sum()) if c in df else 0.0
    row["acos"] = row["cost"] / row["sales7d"] if row["sales7d"] else pd.NA
    row["cpc"] = row["cost"] / row["clicks"] if row["clicks"] else pd.NA
    return row


def build_validation(sources):
    rows = [totals_row(k, v) for k, v in sources.items()]
    df = pd.DataFrame(rows)
    campaign = df[df["source"] == "campaign_summary"].iloc[0]
    checks = []
    for _, row in df.iterrows():
        checks.append({
            "check": f"{row['source']} vs campaign_summary",
            "rows": row["rows"],
            "sales_diff_$": row["sales7d"] - campaign["sales7d"],
            "cost_diff_$": row["cost"] - campaign["cost"],
            "purchase_diff": row["purchases7d"] - campaign["purchases7d"],
            "note": "Different report grains can vary slightly due attribution/reporting scope." if abs(row["sales7d"] - campaign["sales7d"]) > 0.01 else "Matches campaign summary.",
        })
    return df, pd.DataFrame(checks)


def add_existing_checks(add_kw, add_asin, keywords, targets):
    kw = keywords.copy()
    kw["keyword_norm"] = kw["keyword"].map(norm_text)
    exact_enabled = kw[(kw["keywordType"].eq("EXACT")) & (kw["adKeywordStatus"].eq("ENABLED"))]
    exact_map = exact_enabled.groupby("keyword_norm").agg(
        existing_exact_campaigns=("campaignName", base.text_join),
        existing_exact_ad_groups=("adGroupName", base.text_join),
    ).reset_index()
    add_kw = add_kw.copy()
    add_kw["keyword_norm"] = add_kw["searchTerm"].map(norm_text)
    add_kw = add_kw.merge(exact_map, on="keyword_norm", how="left")
    add_kw["already_enabled_exact"] = add_kw["existing_exact_campaigns"].fillna("").ne("")
    add_kw["recommended_action"] = add_kw["already_enabled_exact"].map(
        {True: "AUDIT EXISTING EXACT / MOVE TO WINNERS", False: "ADD NEW EXACT KEYWORD"}
    )
    add_kw["duplicate_check_note"] = add_kw["already_enabled_exact"].map(
        {True: "Exact already exists somewhere; check bid, budget, and serving before adding duplicate.",
         False: "No enabled exact match found in current keyword report."}
    )

    tg = targets.copy()
    tg["target_asin"] = tg["targeting"].map(extract_asin)
    tg["is_exact_asin_target"] = tg["targeting"].astype(str).str.contains('asin="', case=False, regex=False)
    exact_asins = tg[tg["is_exact_asin_target"] & tg["adKeywordStatus"].eq("ENABLED")]
    asin_map = exact_asins.groupby("target_asin").agg(
        existing_exact_asin_campaigns=("campaignName", base.text_join),
        existing_exact_asin_ad_groups=("adGroupName", base.text_join),
    ).reset_index()
    add_asin = add_asin.copy()
    add_asin["target_asin"] = add_asin["searchTerm"].astype(str).str.upper()
    add_asin = add_asin.merge(asin_map, on="target_asin", how="left")
    add_asin["already_enabled_exact_asin"] = add_asin["existing_exact_asin_campaigns"].fillna("").ne("")
    add_asin["recommended_action"] = add_asin["already_enabled_exact_asin"].map(
        {True: "AUDIT EXISTING EXACT ASIN TARGET", False: "ADD / ISOLATE ASIN PRODUCT TARGET"}
    )
    add_asin["duplicate_check_note"] = add_asin["already_enabled_exact_asin"].map(
        {True: "Exact ASIN target already exists somewhere; review serving and bid first.",
         False: "No enabled exact ASIN target found in current target report."}
    )
    return add_kw, add_asin


def build_theme_table(search_agg):
    stop = {
        "the", "and", "for", "with", "adult", "adults", "men", "mens", "man", "women",
        "shirt", "shirts", "tshirt", "tshirts", "tee", "tees", "t", "in", "of", "to",
        "a", "an", "on", "that", "will", "under", "black", "light",
    }
    rows = []
    for _, r in search_agg.iterrows():
        term = norm_text(r["searchTerm"])
        if base.asin_like(term):
            continue
        words = [w for w in re.findall(r"[a-z0-9]+", term) if len(w) > 1 and w not in stop]
        themes = set(words)
        themes.update(" ".join(words[i:i+2]) for i in range(len(words)-1))
        for theme in themes:
            rows.append({
                "theme": theme,
                "searchTerm": r["searchTerm"],
                "impressions": r.get("impressions", 0),
                "clicks": r.get("clicks", 0),
                "cost": r.get("cost", 0),
                "sales7d": r.get("sales7d", 0),
                "purchases7d": r.get("purchases7d", 0),
                "projected_after_ad_profit": r.get("projected_after_ad_profit", 0),
            })
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    out = df.groupby("theme").agg(
        unique_search_terms=("searchTerm", "nunique"),
        example_terms=("searchTerm", lambda x: base.text_join(x, 4)),
        impressions=("impressions", "sum"),
        clicks=("clicks", "sum"),
        cost=("cost", "sum"),
        sales7d=("sales7d", "sum"),
        purchases7d=("purchases7d", "sum"),
        projected_after_ad_profit=("projected_after_ad_profit", "sum"),
    ).reset_index()
    out = enrich_extra(out)
    out["theme_action"] = out.apply(lambda r: (
        "EXPAND / EXACT HARVEST" if r["purchases7d"] >= 2 and (pd.isna(r["acos_calc"]) or r["acos_calc"] <= base.TARGET_ACOS)
        else "NEGATIVE / TIGHTEN REVIEW" if r["purchases7d"] == 0 and r["cost"] >= base.BASE_PRE_AD_PROFIT
        else "WATCH"
    ), axis=1)
    return out.sort_values(["purchases7d", "projected_after_ad_profit", "sales7d"], ascending=False)


def product_tables(advertised, purchased):
    adv = advertised.copy()
    if len(adv):
        adv["other_sku_sales_share"] = adv["salesOtherSku7d"] / adv["sales7d"].replace(0, pd.NA)
        adv["same_sku_sales_share"] = adv["attributedSalesSameSku7d"] / adv["sales7d"].replace(0, pd.NA)
        adv["product_action"] = adv.apply(lambda r: (
            "SCALE SKU ADS" if r["purchases7d"] >= 2 and r["acos_calc"] <= 0.25 and r["projected_after_ad_profit"] > 0
            else "CROSS-SKU LEAKAGE REVIEW" if r["sales7d"] > 0 and r["other_sku_sales_share"] >= 0.80
            else "REDUCE / PAUSE SKU AD" if r["purchases7d"] == 0 and r["cost"] >= base.BASE_PRE_AD_PROFIT
            else "WATCH"
        ), axis=1)
    pur = purchased.copy()
    if len(pur):
        pur["same_as_advertised_asin"] = pur["purchasedAsin"].astype(str).str.upper() == pur["advertisedAsin"].astype(str).str.upper()
        pur["sales_basis"] = pur["salesOtherSku7d"].where(pur["salesOtherSku7d"] > 0, pur["sales7d"])
        pur["purchases_basis"] = pur["purchasesOtherSku7d"].where(pur["purchasesOtherSku7d"] > 0, pur["purchases7d"])
        asin = pur.groupby("purchasedAsin").agg(
            sales_other_sku_7d=("salesOtherSku7d", "sum"),
            purchases_other_sku_7d=("purchasesOtherSku7d", "sum"),
            units_other_sku_7d=("unitsSoldOtherSku7d", "sum"),
            advertised_asins=("advertisedAsin", base.text_join),
            advertised_skus=("advertisedSku", base.text_join),
            campaigns=("campaignName", base.text_join),
            keywords=("keyword", base.text_join),
        ).reset_index()
        asin = asin.sort_values(["purchases_other_sku_7d", "sales_other_sku_7d"], ascending=False)
    else:
        asin = pd.DataFrame()
    return adv, pur, asin


def action_plan(add_kw, add_asin, neg, bid_recs, campaigns, advertised):
    rows = []
    def add(priority, action, logic, evidence, tab):
        rows.append({"priority": priority, "action": action, "logic": logic, "evidence": evidence, "where_to_review": tab})
    if len(add_kw):
        top = add_kw.head(8)["searchTerm"].tolist()
        add(1, "Harvest top converting customer terms into exact/audited exact campaigns",
            "Exact isolates terms already converting from broad/auto/phrase so bids can be controlled.",
            ", ".join(map(str, top)), "Keyword_Adds")
    if len(add_asin):
        top = add_asin.head(8)["searchTerm"].tolist()
        add(2, "Isolate converting ASIN/product targets",
            "ASINs with sales at low ACoS should get exact product targeting or separate bid control.",
            ", ".join(map(str, top)), "ASIN_Target_Adds")
    cuts = bid_recs[bid_recs["recommended_action"].isin(["CUT / PAUSE REVIEW", "REDUCE BID"])].head(8)
    if len(cuts):
        add(3, "Reduce or cut high-waste broad targets first",
            "No-sale spend near one $19.99 profit unit or very high ACoS should not keep eating budget.",
            ", ".join(cuts["targeting"].astype(str).head(8)), "Bid_Recommendations")
    if len(neg):
        add(4, "Add negatives carefully, exact first",
            "Use exact negatives for relevant-but-nonconverting shirt terms; phrase only for clearly wrong product intent.",
            ", ".join(neg["searchTerm"].astype(str).head(8)), "Negative_Review")
    no_serve = campaigns[(campaigns["campaignStatus"].eq("ENABLED")) & (campaigns["impressions"] == 0)]
    if len(no_serve):
        add(5, "Investigate enabled campaigns with zero delivery",
            "Enabled campaigns with zero impressions usually need eligibility, ad group, keyword, bid, or listing checks.",
            ", ".join(no_serve["campaignName"].astype(str)), "Campaign_Summary")
    if len(advertised):
        share = advertised["salesOtherSku7d"].sum() / advertised["sales7d"].sum() if advertised["sales7d"].sum() else 0
        add(6, "Use cross-SKU sales intentionally",
            "Most ad-attributed sales are other-SKU sales, so SKU-level ads may be acting like catalog discovery, not single-design promotion.",
            f"Other-SKU share in advertised product report: {share:.1%}", "Advertised_SKU_Performance")
    return pd.DataFrame(rows)


def main():
    latest_raw, previous_raw, search_agg = base.make_search_term_analysis()
    keywords = base.enrich(base.load_json(f"keywords_latest31_{base.LATEST_START}_{base.LATEST_END}.json")).assign(source_report="keywords")
    targets = base.enrich(base.load_json(f"targets_latest31_{base.LATEST_START}_{base.LATEST_END}.json")).assign(source_report="targets")
    current_targets = pd.concat([keywords, targets], ignore_index=True)
    campaigns = enrich_extra(base.load_json(f"campaigns_latest31_{base.LATEST_START}_{base.LATEST_END}.json"))
    campaigns_daily = load_extra(f"campaigns_daily_latest31_{base.LATEST_START}_{base.LATEST_END}.json")
    placement = load_extra(f"campaign_placement_latest31_{base.LATEST_START}_{base.LATEST_END}.json")
    adgroups = load_extra(f"adgroups_latest31_{base.LATEST_START}_{base.LATEST_END}.json")
    advertised = load_extra(f"advertised_products_latest31_{base.LATEST_START}_{base.LATEST_END}.json")
    purchased = load_extra(f"purchased_products_latest31_{base.LATEST_START}_{base.LATEST_END}.json")

    add_kw, add_asin, neg, bid_recs = base.build_recommendations(search_agg, current_targets)
    add_kw, add_asin = add_existing_checks(add_kw, add_asin, keywords, targets)
    themes = build_theme_table(search_agg)
    advertised, purchased, purchased_asins = product_tables(advertised, purchased)
    action = action_plan(add_kw, add_asin, neg, bid_recs, campaigns, advertised)
    validation_sources = {
        "campaign_summary": campaigns,
        "campaign_daily": campaigns_daily,
        "campaign_placement": placement,
        "adgroups": adgroups,
        "advertised_products": advertised,
        "keyword_targets": keywords,
        "product_targets": targets,
        "search_terms_keyword": base.enrich(base.load_json(f"search_terms_keywords_latest31_{base.LATEST_START}_{base.LATEST_END}.json")),
        "search_terms_target": base.enrich(base.load_json(f"search_terms_targets_latest31_{base.LATEST_START}_{base.LATEST_END}.json")),
    }
    validation_totals, validation_checks = build_validation(validation_sources)

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("README")
    add_title(ws, "Amazon Ads PPC Deep Report", "Read-only analysis. No Amazon Ads account changes, no Google Sheet changes.")
    readme = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Latest window", f"{base.LATEST_START} to {base.LATEST_END}"),
        ("Previous comparison", f"{base.PREV_START} to {base.PREV_END}, search-term reports only"),
        ("Added in this version", "Advertised products, purchased products, daily campaign trend, placement, ad-group, validation checks, theme mining, duplicate exact checks."),
        ("Profit shortcut", f"Projected profit = sales * {base.NET_AFTER_AMAZON_RATE:.0%} - units * ${base.UNIT_COST:.2f} - ad spend."),
        ("Base $19.99 break-even", f"${base.BASE_PRICE:.2f} gives about ${base.BASE_PRE_AD_PROFIT:.2f} pre-ad profit and {base.BASE_BREAKEVEN_ACOS:.1%} break-even ACoS."),
    ]
    for i, (k, v) in enumerate(readme, 4):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, v)
    style_sheet(ws)

    totals = campaigns.sum(numeric_only=True)
    ws = wb.create_sheet("Executive_Summary")
    add_title(ws, "Executive Summary")
    summary = pd.DataFrame([
        ["Spend", totals.get("cost", 0)],
        ["Sales 7d", totals.get("sales7d", 0)],
        ["Purchases 7d", totals.get("purchases7d", 0)],
        ["ACoS", totals.get("cost", 0) / totals.get("sales7d", 1)],
        ["Projected after-ad profit", totals.get("sales7d", 0) * base.NET_AFTER_AMAZON_RATE - totals.get("unitsSoldClicks7d", 0) * base.UNIT_COST - totals.get("cost", 0)],
        ["Keyword add/audit candidates", len(add_kw)],
        ["ASIN target add/audit candidates", len(add_asin)],
        ["Negative review candidates", len(neg)],
        ["Other-SKU sales share", advertised["salesOtherSku7d"].sum() / advertised["sales7d"].sum() if len(advertised) and advertised["sales7d"].sum() else 0],
    ], columns=["metric", "value"])
    write_df(ws, summary, freeze=False, start_row=4)
    insights = [
        "The account is profitable at the latest 31-day level under your 60% net / $5 cost shortcut, but broad targets still contain waste.",
        "The strongest new finding is cross-SKU behavior: advertised product reporting shows most attributed sales come from other SKUs, matching your observation that clicks on one shirt often buy another shirt.",
        "Do not blindly add duplicates: the new keyword and ASIN tabs now flag whether an enabled exact keyword/ASIN already exists, so some rows should be audited or moved rather than duplicated.",
        "The enabled `TSM_SP_MT_KT_EXACT_WINNERS` campaign still shows zero impressions, so it needs a setup/serving check before relying on it as the winner container.",
    ]
    ws["D4"] = "Interpretation"
    ws["D4"].font = Font(bold=True, size=14)
    for i, text in enumerate(insights, 5):
        ws.cell(i, 4, "• " + text)
    style_sheet(ws)
    style_numbers(ws)

    sheets = [
        ("Action_Plan", action),
        ("Validation_Totals", validation_totals),
        ("Validation_Checks", validation_checks),
    ]
    for name, df in sheets:
        ws = wb.create_sheet(name)
        write_df(ws, df)

    add_cols = [
        "recommended_action", "searchTerm", "already_enabled_exact", "existing_exact_campaigns",
        "suggested_match_type", "suggested_start_bid", "purchases7d", "sales7d", "cost", "acos",
        "cpc", "cvr", "projected_after_ad_profit", "prev_purchases7d", "prev_sales7d",
        "campaigns", "matched_keywords", "why", "duplicate_check_note",
    ]
    ws = wb.create_sheet("Keyword_Adds")
    write_df(ws, add_kw[[c for c in add_cols if c in add_kw]].sort_values(["already_enabled_exact", "purchases7d", "sales7d"], ascending=[True, False, False]))

    asin_cols = [
        "recommended_action", "searchTerm", "already_enabled_exact_asin", "existing_exact_asin_campaigns",
        "suggested_start_bid", "purchases7d", "sales7d", "cost", "acos", "cpc",
        "projected_after_ad_profit", "prev_purchases7d", "prev_sales7d", "campaigns",
        "matched_targets", "why", "duplicate_check_note",
    ]
    ws = wb.create_sheet("ASIN_Target_Adds")
    write_df(ws, add_asin[[c for c in asin_cols if c in add_asin]].sort_values(["already_enabled_exact_asin", "purchases7d", "sales7d"], ascending=[True, False, False]))

    neg_cols = [
        "recommended_action", "searchTerm", "clicks", "cost", "impressions", "cpc",
        "prev_purchases7d", "prev_sales7d", "campaigns", "matched_keywords", "matched_targets",
        "non_shirt_signal", "why", "caution",
    ]
    ws = wb.create_sheet("Negative_Review")
    write_df(ws, neg[[c for c in neg_cols if c in neg]].sort_values(["cost", "clicks"], ascending=False))

    bid_cols = [
        "recommended_action", "campaignName", "adGroupName", "targeting", "keyword", "matchType",
        "keywordType", "adKeywordStatus", "keywordBid", "suggested_bid", "impressions", "clicks",
        "cost", "sales7d", "purchases7d", "acos", "cpc", "cvr", "projected_after_ad_profit", "why",
    ]
    ws = wb.create_sheet("Bid_Recommendations_ALL")
    write_df(ws, bid_recs[[c for c in bid_cols if c in bid_recs]])

    campaign_cols = [
        "campaignName", "campaignStatus", "campaignBudgetAmount", "impressions", "clicks", "cost",
        "sales7d", "sales14d", "purchases7d", "purchases14d", "unitsSoldClicks7d", "acos_calc",
        "cpc_calc", "ctr_calc", "cvr_calc", "projected_after_ad_profit",
    ]
    ws = wb.create_sheet("Campaign_Summary")
    write_df(ws, campaigns[[c for c in campaign_cols if c in campaigns]].sort_values("cost", ascending=False))

    ws = wb.create_sheet("Daily_Campaign_Trends")
    daily_cols = ["date", "campaignName", "impressions", "clicks", "cost", "sales7d", "purchases7d", "acos_calc", "topOfSearchImpressionShare", "projected_after_ad_profit"]
    write_df(ws, campaigns_daily[[c for c in daily_cols if c in campaigns_daily]].sort_values(["date", "campaignName"]))
    if ws.max_row > 3:
        chart = LineChart()
        chart.title = "Daily Spend and Sales"
        chart.y_axis.title = "USD"
        chart.x_axis.title = "Rows by date/campaign"
        data = Reference(ws, min_col=5, max_col=6, min_row=1, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.height = 7
        chart.width = 16
        ws.add_chart(chart, "L2")

    ws = wb.create_sheet("Placement_Performance")
    placement_cols = ["campaignName", "placementClassification", "impressions", "clicks", "cost", "sales7d", "purchases7d", "acos_calc", "cpc_calc", "ctr_calc", "cvr_calc", "projected_after_ad_profit"]
    write_df(ws, placement[[c for c in placement_cols if c in placement]].sort_values(["campaignName", "cost"], ascending=[True, False]))
    if ws.max_row > 2:
        chart = BarChart()
        chart.title = "Placement Spend vs Sales"
        chart.y_axis.title = "USD"
        data = Reference(ws, min_col=5, max_col=6, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 16
        ws.add_chart(chart, "N2")

    ws = wb.create_sheet("Ad_Group_Performance")
    adg_cols = ["campaignName", "adGroupName", "adStatus", "impressions", "clicks", "cost", "sales7d", "purchases7d", "acos_calc", "cpc_calc", "ctr_calc", "cvr_calc", "projected_after_ad_profit"]
    write_df(ws, adgroups[[c for c in adg_cols if c in adgroups]].sort_values("cost", ascending=False))

    adv_cols = [
        "product_action", "advertisedSku", "advertisedAsin", "campaignName", "adGroupName",
        "impressions", "clicks", "cost", "sales7d", "purchases7d", "acos_calc",
        "attributedSalesSameSku7d", "salesOtherSku7d", "same_sku_sales_share", "other_sku_sales_share",
        "unitsSoldSameSku7d", "unitsSoldOtherSku7d", "projected_after_ad_profit",
    ]
    ws = wb.create_sheet("Advertised_SKU_Performance")
    write_df(ws, advertised[[c for c in adv_cols if c in advertised]].sort_values(["purchases7d", "sales7d"], ascending=False))

    pur_cols = [
        "purchasedAsin", "sales_other_sku_7d", "purchases_other_sku_7d", "units_other_sku_7d",
        "advertised_asins", "advertised_skus", "campaigns", "keywords",
    ]
    ws = wb.create_sheet("Purchased_ASIN_Summary")
    write_df(ws, purchased_asins[[c for c in pur_cols if c in purchased_asins]].head(1000))

    ws = wb.create_sheet("Purchased_Product_Raw")
    purchased_cols = [
        "campaignName", "adGroupName", "keyword", "keywordType", "matchType", "advertisedSku",
        "advertisedAsin", "purchasedAsin", "same_as_advertised_asin", "salesOtherSku7d",
        "purchasesOtherSku7d", "unitsSoldOtherSku7d", "startDate", "endDate",
    ]
    write_df(ws, purchased[[c for c in purchased_cols if c in purchased]].sort_values(["purchasesOtherSku7d", "salesOtherSku7d"], ascending=False))

    ws = wb.create_sheet("Search_Term_Themes")
    theme_cols = ["theme_action", "theme", "unique_search_terms", "example_terms", "impressions", "clicks", "cost", "sales7d", "purchases7d", "acos_calc", "projected_after_ad_profit"]
    write_df(ws, themes[[c for c in theme_cols if c in themes]].head(1000))

    st_cols = [
        "searchTerm", "is_asin", "non_shirt_signal", "impressions", "clicks", "cost", "sales7d",
        "purchases7d", "unitsSoldClicks7d", "acos", "cpc", "cvr", "projected_after_ad_profit",
        "prev_cost", "prev_sales7d", "prev_purchases7d", "period_sales_delta", "period_cost_delta",
        "campaigns", "matched_keywords", "matched_targets",
    ]
    ws = wb.create_sheet("Search_Term_Analysis")
    write_df(ws, search_agg[[c for c in st_cols if c in search_agg]].sort_values(["purchases7d", "cost"], ascending=[False, False]).head(3000))

    raw_tabs = [
        ("Raw_Search_Terms_Latest", latest_raw),
        ("Raw_Search_Terms_Previous", previous_raw),
        ("Raw_Current_Targets", current_targets),
        ("Raw_Advertised_Products", advertised),
        ("Raw_Campaign_Daily", campaigns_daily),
        ("Raw_Placement", placement),
        ("Raw_AdGroups", adgroups),
    ]
    for name, df in raw_tabs:
        ws = wb.create_sheet(name)
        write_df(ws, df.head(5000))

    ws = wb.create_sheet("Assumptions")
    assumptions = pd.DataFrame([
        ["Amazon net after shortcut", base.NET_AFTER_AMAZON_RATE, "Your simplified all-in Amazon/logistics deduction model."],
        ["Unit cost", base.UNIT_COST, "Flat $5 cost per t-shirt."],
        ["Base price", base.BASE_PRICE, "Reference price for $19.99 pricing strategy."],
        ["Base pre-ad profit", base.BASE_PRE_AD_PROFIT, "Base price * 60% - unit cost."],
        ["Base break-even ACoS", base.BASE_BREAKEVEN_ACOS, "Pre-ad profit / price."],
        ["Target ACoS", base.TARGET_ACOS, "Used as default for scale/keep decisions."],
        ["Negative spend threshold", base.BASE_PRE_AD_PROFIT, "No-sale query spends about one $19.99 unit profit."],
        ["High-click no-sale threshold", 15, "Enough clicks without conversion to review bid/negative."],
    ], columns=["Assumption", "Value", "Meaning"])
    write_df(ws, assumptions)

    wb.save(OUT_FILE)

    # Integrity verification.
    check = load_workbook(OUT_FILE, read_only=True, data_only=False)
    errors = []
    for ws in check.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#"):
                    errors.append((ws.title, cell.coordinate, cell.value))
    print(OUT_FILE)
    print({
        "sheets": len(check.sheetnames),
        "keyword_candidates": len(add_kw),
        "asin_candidates": len(add_asin),
        "negative_candidates": len(neg),
        "advertised_product_rows": len(advertised),
        "purchased_product_rows": len(purchased),
        "theme_rows": len(themes),
        "errors": len(errors),
    })


if __name__ == "__main__":
    main()
