import json
import math
import os
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DATA_DIR = Path("/Users/v/Documents/ART_GPT/outputs/amazon_ads_ppc_20260607")
OUT_FILE = DATA_DIR / "amazon_ads_ppc_keyword_report_2026-06-07.xlsx"

LATEST_START = "2026-05-06"
LATEST_END = "2026-06-05"
PREV_START = "2026-04-05"
PREV_END = "2026-05-05"

NET_AFTER_AMAZON_RATE = 0.60
UNIT_COST = 5.00
BASE_PRICE = 19.99
BASE_PRE_AD_PROFIT = BASE_PRICE * NET_AFTER_AMAZON_RATE - UNIT_COST
BASE_BREAKEVEN_ACOS = BASE_PRE_AD_PROFIT / BASE_PRICE
TARGET_ACOS = 0.35

NON_SHIRT_PATTERNS = [
    "accessor", "bracelet", "necklace", "paint", "makeup", "nail", "poster",
    "decor", "sticker", "vinyl", "lamp", "bulb", "light bulb", "neon sign",
    "party supplies", "glow stick", "glow sticks", "hat", "hoodie", "sweatshirt",
    "dress", "costume", "kids", "toddler", "baby", "girls", "boys",
]


def load_json(name: str) -> pd.DataFrame:
    path = DATA_DIR / name
    if not path.exists():
        return pd.DataFrame()
    with path.open() as f:
        data = json.load(f)
    return pd.DataFrame(data)


def to_num(df: pd.DataFrame, cols) -> pd.DataFrame:
    for col in cols:
        if col not in df.columns:
            df[col] = 0
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def safe_div(a, b):
    return a / b.replace(0, pd.NA)


def enrich(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    numeric = [
        "impressions", "clicks", "cost", "sales7d", "purchases7d",
        "unitsSoldClicks7d", "keywordBid",
    ]
    df = to_num(df, numeric)
    df["ctr"] = safe_div(df["clicks"], df["impressions"])
    df["cpc"] = safe_div(df["cost"], df["clicks"])
    df["cvr"] = safe_div(df["purchases7d"], df["clicks"])
    df["acos"] = safe_div(df["cost"], df["sales7d"])
    df["roas"] = safe_div(df["sales7d"], df["cost"])
    df["avg_order_value"] = safe_div(df["sales7d"], df["purchases7d"])
    units_for_profit = df["unitsSoldClicks7d"].where(df["unitsSoldClicks7d"] > 0, df["purchases7d"])
    df["projected_pre_ad_profit"] = df["sales7d"] * NET_AFTER_AMAZON_RATE - units_for_profit * UNIT_COST
    df["projected_after_ad_profit"] = df["projected_pre_ad_profit"] - df["cost"]
    df["breakeven_acos_est"] = safe_div(df["projected_pre_ad_profit"], df["sales7d"])
    df["breakeven_cpc_est"] = safe_div(df["projected_pre_ad_profit"], df["clicks"])
    return df


def asin_like(value) -> bool:
    s = str(value or "").strip().lower().replace('"', "")
    return bool(re.fullmatch(r"b[0-9a-z]{9}", s))


def looks_nonshirt(value) -> bool:
    s = str(value or "").lower()
    return any(p in s for p in NON_SHIRT_PATTERNS)


def text_join(values, max_items=5):
    out = []
    for v in values:
        if pd.isna(v):
            continue
        s = str(v)
        if s and s not in out:
            out.append(s)
    return ", ".join(out[:max_items])


def money(v):
    if pd.isna(v):
        return ""
    return round(float(v), 2)


def pct(v):
    if pd.isna(v):
        return ""
    return round(float(v), 4)


def nvalue(value, default=0.0):
    if pd.isna(value):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def suggested_bid_from_row(row):
    current = nvalue(row.get("keywordBid"), 0)
    cpc = nvalue(row.get("cpc"), 0)
    be_cpc = nvalue(row.get("breakeven_cpc_est"), 0)
    acos = row.get("acos")
    purchases = nvalue(row.get("purchases7d"), 0)
    cost = nvalue(row.get("cost"), 0)
    clicks = nvalue(row.get("clicks"), 0)
    if current <= 0:
        current = cpc if cpc > 0 else 0.25
    if purchases >= 2 and pd.notna(acos) and acos <= 0.25:
        cap = be_cpc * 0.80 if be_cpc > 0 else current * 1.15
        return round(max(current, min(current * 1.20, cap)), 2)
    if purchases >= 1 and pd.notna(acos) and acos <= TARGET_ACOS:
        cap = be_cpc * 0.75 if be_cpc > 0 else current * 1.05
        return round(max(0.15, min(current * 1.10, cap)), 2)
    if purchases >= 1:
        cap = be_cpc * 0.60 if be_cpc > 0 else current * 0.75
        return round(max(0.05, min(current, current * 0.80, cap)), 2)
    if cost >= BASE_PRE_AD_PROFIT or clicks >= 15:
        return round(max(0.05, current * 0.65), 2)
    if clicks >= 8:
        return round(max(0.05, current * 0.80), 2)
    return round(current, 2)


def target_action(row):
    purchases = nvalue(row.get("purchases7d"), 0)
    clicks = nvalue(row.get("clicks"), 0)
    cost = nvalue(row.get("cost"), 0)
    acos = row.get("acos")
    profit = nvalue(row.get("projected_after_ad_profit"), 0)
    if purchases >= 2 and pd.notna(acos) and acos <= 0.25 and profit > 0:
        return "SCALE +10-20%"
    if purchases >= 1 and pd.notna(acos) and acos <= TARGET_ACOS:
        return "KEEP / LIGHT SCALE"
    if purchases >= 1:
        return "REDUCE BID"
    if cost >= BASE_PRE_AD_PROFIT or clicks >= 15:
        return "CUT / PAUSE REVIEW"
    if clicks >= 8:
        return "REDUCE 15-25%"
    if clicks == 0:
        return "NO CLICK DATA"
    return "KEEP COLLECTING"


def action_reason(row):
    purchases = int(row.get("purchases7d") or 0)
    clicks = int(row.get("clicks") or 0)
    cost = float(row.get("cost") or 0)
    sales = float(row.get("sales7d") or 0)
    acos = row.get("acos")
    profit = float(row.get("projected_after_ad_profit") or 0)
    if purchases:
        a = f"{purchases} purchase(s), ${sales:,.2f} sales"
        if pd.notna(acos):
            a += f", {acos:.1%} ACoS"
        a += f", est. after-ad profit ${profit:,.2f}"
        return a
    return f"No purchases from {clicks} clicks and ${cost:,.2f} spend"


def make_search_term_analysis():
    sk = load_json(f"search_terms_keywords_latest31_{LATEST_START}_{LATEST_END}.json")
    st = load_json(f"search_terms_targets_latest31_{LATEST_START}_{LATEST_END}.json")
    latest = pd.concat(
        [
            sk.assign(source_report="keyword_search_terms"),
            st.assign(source_report="target_search_terms"),
        ],
        ignore_index=True,
    )
    latest = enrich(latest)
    prev = pd.concat(
        [
            load_json(f"search_terms_keywords_previous31_{PREV_START}_{PREV_END}.json").assign(source_report="keyword_search_terms"),
            load_json(f"search_terms_targets_previous31_{PREV_START}_{PREV_END}.json").assign(source_report="target_search_terms"),
        ],
        ignore_index=True,
    )
    prev = enrich(prev) if len(prev) else prev

    agg = latest.groupby("searchTerm", dropna=False).agg(
        impressions=("impressions", "sum"),
        clicks=("clicks", "sum"),
        cost=("cost", "sum"),
        sales7d=("sales7d", "sum"),
        purchases7d=("purchases7d", "sum"),
        unitsSoldClicks7d=("unitsSoldClicks7d", "sum"),
        campaigns=("campaignName", text_join),
        ad_groups=("adGroupName", text_join),
        matched_keywords=("keyword", text_join),
        matched_targets=("targeting", text_join),
        source_reports=("source_report", text_join),
    ).reset_index()
    agg = enrich(agg)
    if len(prev):
        p = prev.groupby("searchTerm", dropna=False).agg(
            prev_impressions=("impressions", "sum"),
            prev_clicks=("clicks", "sum"),
            prev_cost=("cost", "sum"),
            prev_sales7d=("sales7d", "sum"),
            prev_purchases7d=("purchases7d", "sum"),
        ).reset_index()
        agg = agg.merge(p, on="searchTerm", how="left")
    for c in ["prev_impressions", "prev_clicks", "prev_cost", "prev_sales7d", "prev_purchases7d"]:
        if c not in agg:
            agg[c] = 0
        agg[c] = agg[c].fillna(0)
    agg["is_asin"] = agg["searchTerm"].map(asin_like)
    agg["non_shirt_signal"] = agg["searchTerm"].map(looks_nonshirt)
    agg["period_sales_delta"] = agg["sales7d"] - agg["prev_sales7d"]
    agg["period_cost_delta"] = agg["cost"] - agg["prev_cost"]
    agg["recommendation_reason"] = agg.apply(action_reason, axis=1)
    return latest, prev, agg


def build_recommendations(search_agg, targets):
    add_kw = search_agg[
        (~search_agg["is_asin"])
        & (search_agg["purchases7d"] >= 1)
        & (
            (search_agg["acos"].fillna(999) <= TARGET_ACOS)
            | ((search_agg["purchases7d"] >= 2) & (search_agg["acos"].fillna(999) <= 0.50))
        )
    ].copy()
    add_kw["recommended_action"] = "ADD EXACT KEYWORD"
    add_kw["suggested_match_type"] = "EXACT"
    add_kw["suggested_start_bid"] = add_kw.apply(lambda r: round(min(max(float(r.get("cpc") or 0.30) * 1.10, 0.25), max(float(r.get("breakeven_cpc_est") or 0.65) * 0.75, 0.25), 0.85), 2), axis=1)
    add_kw["why"] = add_kw.apply(action_reason, axis=1)
    add_kw = add_kw.sort_values(["purchases7d", "projected_after_ad_profit", "sales7d"], ascending=False)

    add_asin = search_agg[
        (search_agg["is_asin"])
        & (search_agg["purchases7d"] >= 1)
        & (search_agg["acos"].fillna(999) <= TARGET_ACOS)
    ].copy()
    add_asin["recommended_action"] = "ADD / ISOLATE ASIN PRODUCT TARGET"
    add_asin["suggested_start_bid"] = add_asin.apply(lambda r: round(min(max(float(r.get("cpc") or 0.30) * 1.05, 0.20), max(float(r.get("breakeven_cpc_est") or 0.55) * 0.70, 0.20), 0.75), 2), axis=1)
    add_asin["why"] = add_asin.apply(action_reason, axis=1)
    add_asin = add_asin.sort_values(["purchases7d", "projected_after_ad_profit", "sales7d"], ascending=False)

    neg = search_agg[
        (search_agg["purchases7d"] == 0)
        & (
            (search_agg["cost"] >= BASE_PRE_AD_PROFIT)
            | (search_agg["clicks"] >= 15)
            | ((search_agg["non_shirt_signal"]) & (search_agg["cost"] >= 2.50))
        )
    ].copy()
    neg["recommended_action"] = neg.apply(
        lambda r: "NEGATIVE PRODUCT/ASIN REVIEW" if r["is_asin"]
        else ("NEGATIVE PHRASE REVIEW" if r["non_shirt_signal"] else "NEGATIVE EXACT REVIEW"),
        axis=1,
    )
    neg["why"] = neg.apply(action_reason, axis=1)
    neg["caution"] = neg.apply(
        lambda r: "Phrase negative only if this is truly outside your catalog." if r["non_shirt_signal"] and not r["is_asin"]
        else "Exact negative is safer than phrase negative for relevant shirt terms.",
        axis=1,
    )
    neg = neg.sort_values(["cost", "clicks"], ascending=False)

    t = targets.copy()
    t["recommended_action"] = t.apply(target_action, axis=1)
    t["suggested_bid"] = t.apply(suggested_bid_from_row, axis=1)
    t["why"] = t.apply(action_reason, axis=1)
    t = t.sort_values(["projected_after_ad_profit", "purchases7d", "cost"], ascending=[True, False, False])
    return add_kw, add_asin, neg, t


def build_campaign_summary():
    df = enrich(load_json(f"campaigns_latest31_{LATEST_START}_{LATEST_END}.json"))
    if len(df):
        df["recommended_note"] = df.apply(
            lambda r: "No data yet; campaign may be new, not serving, or blocked by setup/budget/eligibility." if r["clicks"] == 0
            else ("Strong efficiency; protect budget and mine search terms." if r["acos"] <= 0.20 else ("Profitable but watch bids/search waste." if r["acos"] <= TARGET_ACOS else "Above target; reduce waste/bids.")),
            axis=1,
        )
    return df


def write_df(ws, df, start_row=1, start_col=1, title=None, freeze=True, max_rows=None):
    r = start_row
    if title:
        ws.cell(r, start_col, title).font = Font(bold=True, size=14, color="FFFFFF")
        ws.cell(r, start_col).fill = PatternFill("solid", fgColor="111827")
        r += 1
    use = df.head(max_rows).copy() if max_rows else df.copy()
    # Convert NaNs/NA to blanks.
    use = use.where(pd.notna(use), "")
    for j, col in enumerate(use.columns, start_col):
        c = ws.cell(r, j, str(col))
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for i, row in enumerate(use.itertuples(index=False), r + 1):
        for j, value in enumerate(row, start_col):
            if isinstance(value, float) and (math.isinf(value) or math.isnan(value)):
                value = ""
            ws.cell(i, j, value)
    last_row = r + len(use)
    last_col = start_col + len(use.columns) - 1
    if freeze:
        ws.freeze_panes = ws.cell(r + 1, start_col).coordinate
    if len(use.columns):
        ws.auto_filter.ref = f"{get_column_letter(start_col)}{r}:{get_column_letter(last_col)}{last_row}"
    format_sheet(ws)
    return last_row, last_col


def format_sheet(ws):
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
            if isinstance(cell.value, (int, float)):
                header = str(ws.cell(1 if ws.max_row == 1 else 1, cell.column).value or "").lower()
    for col in range(1, min(ws.max_column, 30) + 1):
        max_len = 8
        for row in range(1, min(ws.max_row, 1000) + 1):
            value = ws.cell(row, col).value
            if value is None:
                continue
            max_len = max(max_len, min(len(str(value)), 55))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 10), 42)


def style_numeric(ws):
    percent_names = {"ctr", "cvr", "acos", "breakeven_acos_est"}
    money_names = {"cost", "sales7d", "cpc", "keywordBid", "suggested_bid", "suggested_start_bid", "projected_pre_ad_profit", "projected_after_ad_profit", "avg_order_value", "breakeven_cpc_est", "prev_cost", "prev_sales7d", "period_sales_delta", "period_cost_delta", "campaignBudgetAmount"}
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for name, col in headers.items():
        if name in percent_names:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).number_format = "0.0%"
        if name in money_names:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).number_format = '$#,##0.00'
    if "acos" in headers:
        col = headers["acos"]
        rng = f"{get_column_letter(col)}2:{get_column_letter(col)}{ws.max_row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThanOrEqual", formula=["0.25"], fill=PatternFill("solid", fgColor="C6EFCE")))
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=[str(TARGET_ACOS)], fill=PatternFill("solid", fgColor="FFC7CE")))


def compact_cols(df, cols):
    return df[[c for c in cols if c in df.columns]].copy()


def main():
    latest_raw, previous_raw, search_agg = make_search_term_analysis()
    keywords = enrich(load_json(f"keywords_latest31_{LATEST_START}_{LATEST_END}.json")).assign(source_report="keywords")
    targets = enrich(load_json(f"targets_latest31_{LATEST_START}_{LATEST_END}.json")).assign(source_report="targets")
    current_targets = pd.concat([keywords, targets], ignore_index=True)
    campaigns = build_campaign_summary()
    add_kw, add_asin, neg, bid_recs = build_recommendations(search_agg, current_targets)

    totals = {
        "Spend": campaigns["cost"].sum(),
        "Sales": campaigns["sales7d"].sum(),
        "Purchases": campaigns["purchases7d"].sum(),
        "Units": campaigns["unitsSoldClicks7d"].sum(),
        "Impressions": campaigns["impressions"].sum(),
        "Clicks": campaigns["clicks"].sum(),
    }
    totals["ACoS"] = totals["Spend"] / totals["Sales"] if totals["Sales"] else 0
    totals["CTR"] = totals["Clicks"] / totals["Impressions"] if totals["Impressions"] else 0
    totals["CPC"] = totals["Spend"] / totals["Clicks"] if totals["Clicks"] else 0
    totals["CVR"] = totals["Purchases"] / totals["Clicks"] if totals["Clicks"] else 0
    totals["Projected pre-ad profit"] = totals["Sales"] * NET_AFTER_AMAZON_RATE - totals["Units"] * UNIT_COST
    totals["Projected after-ad profit"] = totals["Projected pre-ad profit"] - totals["Spend"]

    prev_total = previous_raw[["cost", "sales7d", "purchases7d", "clicks"]].sum(numeric_only=True) if len(previous_raw) else pd.Series(dtype=float)
    latest_total = latest_raw[["cost", "sales7d", "purchases7d", "clicks"]].sum(numeric_only=True)

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("README")
    ws["A1"] = "Amazon Ads PPC Keyword Report"
    ws["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="111827")
    notes = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Scope", "Sponsored Products, US seller profile, read-only Amazon Ads report exports."),
        ("Latest window", f"{LATEST_START} to {LATEST_END}"),
        ("Previous comparison", f"{PREV_START} to {PREV_END}, search-term reports only"),
        ("No account changes", "No bids, campaigns, keywords, negatives, or budgets were changed."),
        ("Profit shortcut", f"Projected pre-ad profit = sales * {NET_AFTER_AMAZON_RATE:.0%} - units * ${UNIT_COST:.2f}."),
        ("$19.99 base math", f"${BASE_PRICE:.2f} * 60% - ${UNIT_COST:.2f} = ${BASE_PRE_AD_PROFIT:.2f} pre-ad profit; break-even ACoS about {BASE_BREAKEVEN_ACOS:.1%}."),
        ("Decision rule", "Scale low-ACoS converting terms, harvest converting broad/auto terms into exact/product targets, and review high-click no-sale terms for negative exact/phrase."),
    ]
    for i, (k, v) in enumerate(notes, 3):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, v)
    format_sheet(ws)

    ws = wb.create_sheet("Executive_Summary")
    ws["A1"] = "Executive Summary"
    ws["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="111827")
    summary_rows = [
        ("Spend", totals["Spend"]),
        ("Sales", totals["Sales"]),
        ("Purchases", totals["Purchases"]),
        ("ACoS", totals["ACoS"]),
        ("CPC", totals["CPC"]),
        ("CTR", totals["CTR"]),
        ("CVR", totals["CVR"]),
        ("Projected pre-ad profit", totals["Projected pre-ad profit"]),
        ("Projected after-ad profit", totals["Projected after-ad profit"]),
        ("Exact keyword adds found", len(add_kw)),
        ("Product/ASIN target adds found", len(add_asin)),
        ("Negative review candidates", len(neg)),
    ]
    for r, (metric, val) in enumerate(summary_rows, 3):
        ws.cell(r, 1, metric).font = Font(bold=True)
        ws.cell(r, 2, val)
        if metric in {"Spend", "Sales", "CPC", "Projected pre-ad profit", "Projected after-ad profit"}:
            ws.cell(r, 2).number_format = '$#,##0.00'
        if metric in {"ACoS", "CTR", "CVR"}:
            ws.cell(r, 2).number_format = "0.0%"
    ws["D3"] = "What this means"
    ws["D3"].font = Font(bold=True, size=14)
    insights = [
        f"Latest Sponsored Products performance is profitable under your shortcut margin model: {totals['ACoS']:.1%} ACoS versus about {BASE_BREAKEVEN_ACOS:.1%} break-even at $19.99.",
        f"Latest search-term sales were ${latest_total.get('sales7d', 0):,.2f} on ${latest_total.get('cost', 0):,.2f} spend. Previous search-term window was ${prev_total.get('sales7d', 0):,.2f} on ${prev_total.get('cost', 0):,.2f} spend.",
        "The main opportunity is harvesting converting broad/auto terms into exact keywords and isolated ASIN targets, then giving proven terms controlled bid room.",
        "The main risk is broad relevant-but-nonconverting traffic. Use exact negatives first; phrase negatives only for clearly wrong product intent like accessories, paint, bulbs, decor, or kids items.",
    ]
    for i, text in enumerate(insights, 4):
        ws.cell(i, 4, "• " + text)
    format_sheet(ws)

    cols_add = ["recommended_action", "searchTerm", "suggested_match_type", "suggested_start_bid", "purchases7d", "sales7d", "cost", "acos", "cpc", "cvr", "projected_after_ad_profit", "campaigns", "matched_keywords", "why"]
    ws = wb.create_sheet("Keyword_Adds")
    write_df(ws, compact_cols(add_kw, cols_add), max_rows=300)
    style_numeric(ws)

    cols_asin = ["recommended_action", "searchTerm", "suggested_start_bid", "purchases7d", "sales7d", "cost", "acos", "cpc", "projected_after_ad_profit", "campaigns", "matched_targets", "why"]
    ws = wb.create_sheet("ASIN_Target_Adds")
    write_df(ws, compact_cols(add_asin, cols_asin), max_rows=200)
    style_numeric(ws)

    cols_neg = ["recommended_action", "searchTerm", "clicks", "cost", "impressions", "cpc", "campaigns", "matched_keywords", "matched_targets", "non_shirt_signal", "why", "caution"]
    ws = wb.create_sheet("Negative_Review")
    write_df(ws, compact_cols(neg, cols_neg), max_rows=300)
    style_numeric(ws)

    cols_bid = ["recommended_action", "campaignName", "adGroupName", "targeting", "keyword", "matchType", "keywordType", "adKeywordStatus", "keywordBid", "suggested_bid", "impressions", "clicks", "cost", "sales7d", "purchases7d", "acos", "cpc", "cvr", "projected_after_ad_profit", "why"]
    ws = wb.create_sheet("Bid_Recommendations")
    write_df(ws, compact_cols(bid_recs, cols_bid), max_rows=600)
    style_numeric(ws)

    cols_campaign = ["campaignName", "campaignStatus", "campaignBudgetAmount", "impressions", "clicks", "cost", "sales7d", "purchases7d", "unitsSoldClicks7d", "acos", "cpc", "ctr", "cvr", "projected_after_ad_profit", "recommended_note"]
    ws = wb.create_sheet("Campaign_Summary")
    write_df(ws, compact_cols(campaigns.sort_values("cost", ascending=False), cols_campaign))
    style_numeric(ws)
    if ws.max_row >= 3:
        chart = BarChart()
        chart.title = "Spend vs Sales by Campaign"
        chart.y_axis.title = "USD"
        chart.x_axis.title = "Campaign"
        data = Reference(ws, min_col=6, max_col=7, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 16
        ws.add_chart(chart, "Q2")

    cols_st = ["searchTerm", "is_asin", "non_shirt_signal", "impressions", "clicks", "cost", "sales7d", "purchases7d", "unitsSoldClicks7d", "acos", "cpc", "cvr", "projected_after_ad_profit", "prev_cost", "prev_sales7d", "prev_purchases7d", "period_sales_delta", "period_cost_delta", "campaigns", "matched_keywords", "matched_targets"]
    ws = wb.create_sheet("Search_Term_Analysis")
    write_df(ws, compact_cols(search_agg.sort_values(["purchases7d", "cost"], ascending=[False, False]), cols_st), max_rows=2000)
    style_numeric(ws)

    ws = wb.create_sheet("Raw_Search_Terms_Latest")
    raw_cols = ["source_report", "campaignName", "adGroupName", "targeting", "keyword", "matchType", "keywordType", "searchTerm", "impressions", "clicks", "cost", "sales7d", "purchases7d", "unitsSoldClicks7d", "startDate", "endDate"]
    write_df(ws, compact_cols(latest_raw, raw_cols), max_rows=5000)
    style_numeric(ws)

    ws = wb.create_sheet("Raw_Current_Targets")
    write_df(ws, compact_cols(current_targets, cols_bid[:-2]), max_rows=5000)
    style_numeric(ws)

    ws = wb.create_sheet("Raw_Search_Terms_Previous")
    write_df(ws, compact_cols(previous_raw, raw_cols), max_rows=5000)
    style_numeric(ws)

    ws = wb.create_sheet("Assumptions")
    rows = [
        ("Amazon net after shortcut", NET_AFTER_AMAZON_RATE, "This is your simplified all-in Amazon/logistics deduction model."),
        ("Unit cost", UNIT_COST, "Flat $5 cost per t-shirt."),
        ("Base price", BASE_PRICE, "Useful reference if testing everything at $19.99."),
        ("Base pre-ad profit", BASE_PRE_AD_PROFIT, "Base price * net rate - unit cost."),
        ("Base break-even ACoS", BASE_BREAKEVEN_ACOS, "Pre-ad profit / price."),
        ("Target ACoS used for recommendations", TARGET_ACOS, "Conservative default for $19.99 products."),
        ("Negative exact waste threshold", BASE_PRE_AD_PROFIT, "If a no-sale query spends roughly one $19.99 unit profit, review it."),
        ("High-click no-sale threshold", 15, "Enough clicks without conversion to review exact negative or bid reduction."),
    ]
    write_df(ws, pd.DataFrame(rows, columns=["Assumption", "Value", "Meaning"]))
    ws["B2"].number_format = "0%"
    ws["B6"].number_format = "0.0%"
    for r in [3, 4, 5, 7]:
        ws.cell(r, 2).number_format = '$#,##0.00'

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="1F4E78")
        format_sheet(ws)

    wb.save(OUT_FILE)
    print(OUT_FILE)
    print(json.dumps({
        "latest_spend": round(totals["Spend"], 2),
        "latest_sales": round(totals["Sales"], 2),
        "latest_purchases": int(totals["Purchases"]),
        "latest_acos": round(totals["ACoS"], 4),
        "keyword_adds": int(len(add_kw)),
        "asin_target_adds": int(len(add_asin)),
        "negative_review": int(len(neg)),
    }, indent=2))


if __name__ == "__main__":
    main()
